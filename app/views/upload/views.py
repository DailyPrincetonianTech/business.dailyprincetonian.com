from flask import Blueprint, request, render_template, redirect, current_app
from openpyxl import load_workbook

from app.database import db
from app.services.advertisement import AdvertisementObjectOptions, create_advertisements, delete_all_advertisements



blueprint = Blueprint("upload", __name__)

# GET/POST /upload
@blueprint.route("/upload", methods = ["GET", "POST"])
def upload():
    
    # If GET, then render the upload page.
    if request.method == "GET":
        return render_template("/upload/index.html")
    
    
    # If POST, then check authorization credentials.
    if request.method == "POST":
        if request.form.get("upload_auth_token") == current_app.config["UPLOAD_AUTH_TOKEN"]:
            # If authorized, then validate and process the file.
            file = request.files.get("file")
            if not file:
                return render_template("/upload/index.html", response = "No file was uploaded.")
            
            errors = []
            if _validateFile(file, errors) and _processFile(file, errors):    
                return render_template("/upload/index.html", response = "File uploaded successfully.")
            # File did not pass validation or processing.
            return render_template("/upload/index.html", response = "File upload failed.", errors = errors)

        # Authorization failed.
        return render_template("/upload/index.html", response = "Invalid authorization token.")
    
    
    
def _validateFile(file, errors):
    VALID_EXTENSIONS = ["XSL", "XLSX"]
    
    # Check if file has a valid extension.
    if file.filename.split(".")[-1].upper() not in VALID_EXTENSIONS:
        errors.append("Invalid file extension.")
        return False
    
    # File is valid.
    return True
    
    
    
def _processFile(file, errors):
    COLUMN_NAMES = ["audience_id", "advertisement_title", "options", "asterisks", "popup", "image_url"]
    worksheets = load_workbook(file, data_only = True).worksheets
    advertisement_dtos = []
    
    # Delete all advertisements.
    # Note that this is not committed yet.
    delete_errors = delete_all_advertisements()
    for error in delete_errors:
        errors.append(error)

    # Process each worksheet.
    for worksheet in worksheets:
        first_row = True
        for row in worksheet.iter_rows():
            # Skip first row as it is the column names,
            # or if the row is empty.
            if first_row or row[0].value is None:
                first_row = False
                continue
            
            # Compile row data.
            values = {}
            for cell in row:
                if cell.value is not None:
                    values[COLUMN_NAMES[cell.column - 1]] = cell.value
            
            # Create objects from compiled data.
            try:
                advertisement_dtos.append(AdvertisementObjectOptions.from_excel_schema(**values))
            except Exception as e:
                errors.append("Row " + str(row[0].row) + ": " + str(e))
        
    create_errors = create_advertisements(advertisement_dtos)
    for error in create_errors:
        errors.append(error)


    # If there are no errors, commit changes to database.
    if not errors:
        db.session.commit()
        return True
    else:
        return False